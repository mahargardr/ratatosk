import pandas as pd
import numpy as np
import re
import ast
import os
from copy import copy
import string
from typing import Dict, Optional

from pandas.api.types import is_numeric_dtype
from .config_reference import ConfigReference
from .cell_list import CellList
from ratatosk.global_config import GlobalConfig

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Fill, PatternFill, Alignment, Border, Side, Font
from openpyxl.styles.colors import Color

class auditResult:
    def __init__(
            self,
            dict_audit: Dict[str, pd.DataFrame],
            config_reference: ConfigReference
    ):
        self.audit_result = dict_audit
        self.config_reference   = config_reference

    def create_report(
            self,
            file_path: str,
            ran_objects: CellList,
            dict_cm: Optional[list],
            verbose: bool = False,
    ) -> None:
        if not verbose:
            df = self.__create_simple_report()
            df.to_excel(file_path,index=None)
            print('  Done')
        else:
            df = self.__create_verbose_report(file_path,
                                              ran_objects,
                                              dict_cm)
        
        return None

    def __create_simple_report(
            self
    ) -> pd.DataFrame :
        print('Creating simple summary.')
        df_brief_summary = pd.DataFrame(columns=['mecontext','MO Class','MO','Parameter','Value','Recommended Setting','Status'])
        df_format = self.config_reference.settings

        for sub_mo in self.audit_result:
            print(sub_mo)

            mo_id = sub_mo.split('=')[0].lower()+'id'

            df_concheck = self.audit_result[sub_mo]

            df_concheck['MO'] = ''
            sub_ids = ['eutrancellfddid','eutrancelltddid','nrcellduid',mo_id]
            for col in df_concheck.columns:
                if (col.endswith('id')) and (col in sub_ids):
                    # print(col)
                    # print(pd.unique(df_concheck[col]))
                    df_concheck.loc[df_concheck['MO']!='','MO'] = df_concheck['MO']+','
                    df_concheck['MO'] = df_concheck['MO']+'%s='%col.replace('id','')
                    df_concheck['MO'] = df_concheck['MO']+df_concheck[col].fillna('').astype(str).str.replace('\.0','')

            parameters = list(pd.unique(self.config_reference.settings.loc[self.config_reference.settings['MO']==sub_mo]['Parameter']))

            #print(df_concheck)

            for param in parameters:
                param = param.lower()
                df_concheck_summary = df_concheck[['mecontext','MO',param,param+'_ref',param+'_check']]
                df_concheck_summary['Parameter'] = param
                df_concheck_summary['MO Class'] = sub_mo.split('=')[0]
                df_concheck_summary = df_concheck_summary.rename(columns={param : 'Value',
                                                                          param+'_ref':'Recommended Setting',
                                                                          param+'_check':'Status'})
                df_concheck_summary = df_concheck_summary[['mecontext',
                                                           'MO Class',
                                                           'MO',
                                                           'Parameter',
                                                           'Value',
                                                           'Recommended Setting',
                                                           'Status'
                                                           ]]
                df_brief_summary = pd.concat([df_brief_summary,df_concheck_summary])

        return df_brief_summary

    def __create_verbose_report(
            self,
            file_path: str,
            ran_objects: pd.DataFrame,
            dict_df: Optional[Dict[str,pd.DataFrame]]
    ) -> None :
        print('\n-----------------------------')
        print('Creating Verbose Report')
        print('-----------------------------')

        wb = Workbook()

        df_format = self.config_reference.settings
        df_sheet = ran_objects[['mecontext', 'siteid', 'cell', 'dlChannelBandwidth']]
        df_sheet['duplex_type'] = np.nan
        dict_concheck = self.audit_result
        config_reference = self.config_reference
        common_col = df_sheet.columns

        dict_sheet = {}
        summary_data = []
        summary_index = [2]
        group_summary_index = [3]

        new_sub_mo_rows = []

        for indicator in pd.unique(df_format['Parameter Indicator']):
            dict_sheet[indicator] = wb.create_sheet(indicator)
            print("%s" %indicator)

            indicator_col = []
            indicator_check_col = []

            df_indicator = df_format.loc[df_format['Parameter Indicator'] == indicator]
            df_sheet['check_overall_'+indicator] = np.nan

            indicator_pos = 5
            color_index = 40
            param_start_index = [indicator_pos]
            group_letter =[]

            indicator_summary_dict = {'Parameter Indicator': indicator, 'OK': 0, 'NOK': 0, 'NA': ''}
            group_summary_list = []

            for group in pd.unique(df_indicator['Group Parameter']):
                print("       %s" %group)
                
                group_member_count = 0
                group_color = PatternFill(patternType='solid', fgColor=Color(indexed=color_index))

                df_group = df_indicator.loc[df_indicator['Group Parameter'] == group]

                df_sheet['check_'+group] = np.nan
                d = dict_sheet[indicator].cell(row=1, column=indicator_pos, value=group)

                param_summary_list = []

                for index,row in df_group.iterrows():
                    param = row['Parameter']
                    mo = row['MO']
                    action = row['Action']
                    remark = row['Remark']
                    rules = row['Rules']

                    if ('CellRelation' in mo) or ('External' in mo):
                        continue

                    index_col = ['mecontext']
                    cell_cols = ['eutrancellfddid','eutrancelltddid','nrcellduid']
                    for col in cell_cols:
                        if (col in dict_concheck[mo].columns) & (col not in index_col):
                            index_col.append('cell')
                    
                    sub_ids = []
                    df_concheck = dict_concheck[mo].copy()
                    mo_id = mo.lower()+'id'
                    if ("=*" in mo):
                        if (mo_id not in cell_cols) & (mo_id in df_concheck.columns):
                            sub_ids = list(pd.unique(df_concheck[mo_id]))
                        else:
                            sub_ids=['']
                    else:
                        sub_ids=['']
                    
                    # merge cells
                    df_concheck['cell'] = ''
                    df_concheck['duplex_type'] = ''
                    for cell_col in cell_cols:
                        #print(cell_col)
                        if cell_col in df_concheck.columns:
                            df_concheck.loc[(pd.notnull(df_concheck[cell_col]) & (df_concheck['cell']=='')), 'cell'] = df_concheck[cell_col]
                            df_concheck.loc[(pd.notnull(df_concheck[cell_col]) & (df_concheck['duplex_type']=='')), 'duplex_type'] = cell_col.replace('id','')
                            #print(f"df concheck : {df_concheck}")
                    for sub_id in sub_ids:
                        if sub_id != '':
                            df_sub_concheck = df_concheck.loc[df_concheck[mo_id]==sub_id]
                            sub_mo = mo+"="+str(sub_id)
                            new_sub_mo_rows.append({
                                'MO.Parameter'  : sub_mo+'.'+param,
                                'Action'        : action,
                                'Remark'        : remark,
                                'Rules'         : rules
                            })
                        else:
                            df_sub_concheck = df_concheck
                            sub_mo=mo

                        #print(f"dfconcheck: {df_sub_concheck}")

                        col_name = sub_mo+'.'+param

                        dict_concheck_mapper = df_sub_concheck.set_index(index_col).to_dict()

                        df_sheet[col_name] = np.nan
                        df_sheet.loc[pd.isnull(df_sheet[col_name]),col_name] = df_sheet.set_index(index_col).index.map(dict_concheck_mapper[param])
                        
                        #print(index_col)
                        
                        if mo.lower()+'id' in cell_cols:
                            df_sheet.loc[pd.isnull(df_sheet['duplex_type']),'duplex_type'] = df_sheet.loc[pd.isnull(df_sheet['duplex_type'])].set_index(index_col).index.map(dict_concheck_mapper['duplex_type'])
                            df_sheet.loc[(pd.isnull(df_sheet[col_name]) & (df_sheet['duplex_type']!=mo.lower())) ,col_name] = "-"
                            #print(f"df sheet : {df_sheet.loc[(pd.isnull(df_sheet[col_name]) & (df_sheet['duplex_type']!=mo.lower()))]}")
                            #if 'eutrancellfdd' in col_name.lower():
                            #print(f"df sheet : {df_sheet}")

                        indicator_col.append(col_name)
                        group_member_count+=1

                        df_sheet[col_name+'_ref'] = np.nan
                        df_sheet[col_name+'_check'] = np.nan
                        try:
                            df_sheet.loc[pd.isnull(df_sheet[col_name+'_ref']),col_name+'_ref'] = df_sheet.set_index(index_col).index.map(dict_concheck_mapper[param+'_ref'])
                            df_sheet.loc[pd.isnull(df_sheet[col_name+'_check']),col_name+'_check'] = df_sheet.set_index(index_col).index.map(dict_concheck_mapper[param+'_check'])

                            indicator_col.append(col_name+'_ref')
                            indicator_col.append(col_name+'_check')

                            group_member_count+=2

                        except Exception as err:
                            print(err)
                            df_sheet[col_name+'_check'] = np.nan
                            indicator_col.append(col_name+'_check')
                            group_member_count+=1

                        print("          %s" %col_name)
                        
                        if action.lower() == 'as info':
                            df_sheet.loc[pd.notnull(df_sheet[col_name]),col_name+'_check'] = 'OK'
                            df_sheet[col_name+'_ref'] = rules
                        
                        df_sheet.loc[((pd.isnull(df_sheet[col_name+'_check'])) & (df_sheet[col_name]=="-")),col_name+'_check'] = 'OK'

                        #df_sheet.loc[(pd.isnull(df_sheet[col_name+'_check'])) | (pd.isnull(df_sheet[col_name])),'check_'+group] = 'NA'
                        df_sheet.loc[(pd.isnull(df_sheet['check_'+group])) & (df_sheet[col_name+'_check'] == 'NOK'),'check_'+group] = 'NOK'

                        summary_dict = self.create_summary_dict(df_sheet[col_name+'_check'])
                        param_summary = {'Parameter Indicator': col_name, 
                                        'OK': summary_dict['OK'], 
                                        'NOK': summary_dict['NOK'], 
                                        'NA': summary_dict['NA']}
                        param_summary_list.append(param_summary)

                df_sheet.loc[(pd.isnull(df_sheet['check_'+group])),'check_'+group] = 'OK'

                #df_sheet.loc[df_sheet['check_'+group]=='NA', 'check_overall_'+indicator] = 'NA'
                df_sheet.loc[(pd.isnull(df_sheet['check_overall_'+indicator])) & (df_sheet['check_'+group] == 'NOK'),'check_overall_'+indicator] = 'NOK'

                summary_dict = self.create_summary_dict(df_sheet['check_'+group])
                group_summary = {'Parameter Indicator': group, 
                                'OK': summary_dict['OK'], 
                                'NOK': summary_dict['NOK'], 
                                'NA': summary_dict['NA']}
                group_summary_list.append(group_summary)
                group_summary_list += param_summary_list

                dict_sheet[indicator].merge_cells(start_row=1,start_column=indicator_pos, end_row=1 , end_column=indicator_pos+group_member_count-1)
                d.alignment = Alignment(horizontal="center",vertical='center')
                d.fill = group_color

                indicator_pos+=group_member_count
                indicator_check_col.append('check_'+group)
                param_start_index.append(indicator_pos)
                group_letter.append(d.column_letter)

                color_index += 1

                group_summary_index.append(group_summary_index[-1]+1+int(group_member_count/3))


            df_sheet.loc[(pd.isnull(df_sheet['check_overall_'+indicator])),'check_overall_'+indicator] = 'OK'

            ####### Overall Check Column Formatting ######

            summary_index.append(group_summary_index[-1])
            group_summary_index[-1] += 1

            d = dict_sheet[indicator].cell(row=1, column=indicator_pos, value='Overall Check')
            dict_sheet[indicator].merge_cells(start_row=1,
                                            end_row=1,
                                            start_column=indicator_pos, 
                                            end_column=indicator_pos+len(pd.unique(df_indicator['Group Parameter']))-1)
            d.alignment = Alignment(horizontal="center",vertical="center")
            d.fill = PatternFill(patternType='solid', fgColor=Color(indexed=color_index+1))
            group_letter.append(d.column_letter)

            indicator_check_col.append('check_overall_'+indicator)

            summary_dict = self.create_summary_dict(df_sheet['check_overall_'+indicator])
            indicator_summary_dict['OK'] = summary_dict['OK']
            indicator_summary_dict['NOK'] = summary_dict['NOK']
            indicator_summary_dict['NA'] = summary_dict['NA']
            group_summary_list.insert(0,indicator_summary_dict)

            summary_data += group_summary_list

            presentation_col = list(common_col) + indicator_col + indicator_check_col
            df = df_sheet[presentation_col]
            df = df.drop('duplex_type', axis=1)

            for r in dataframe_to_rows(df,index=None):
                dict_sheet[indicator].append(r)

            for i in range(len(param_start_index)):
                start_i = param_start_index[i]-1
                last_col = False
                if (i == len(param_start_index)-1):
                    stop_i = start_i+len(pd.unique(df_indicator['Group Parameter']))
                    last_col = True
                else:
                    stop_i = param_start_index[i+1]-1

                j=0
                for col in range(start_i,stop_i):
                    try:
                        letter = dict_sheet[indicator][2][col].column_letter
                        if last_col:
                            dict_sheet[indicator][2][col].fill = copy(dict_sheet[indicator][2][param_start_index[j]-1].fill)
                            j+=1
                        else:
                            dict_sheet[indicator][2][col].fill = copy(dict_sheet[indicator][1][start_i].fill)

                        dict_sheet[indicator][2][col].alignment = Alignment(textRotation=45,horizontal="center")
                        
                        header_text = v
                        dict_sheet[indicator].column_dimensions[letter].width = max(len(header_text))
                        
                        if ('ref' in dict_sheet[indicator][2][col].value) or ('_check' in dict_sheet[indicator][2][col].value):     
                            dict_sheet[indicator].column_dimensions[letter].hidden = True
                    except:
                        pass

            dsum = dict_sheet[indicator].cell(row=2, column=indicator_pos+len(pd.unique(df_indicator['Group Parameter'])), value='Summary')
            dsum.alignment = Alignment(textRotation=45,horizontal="center")
            dsum.fill = PatternFill(patternType='solid', fgColor=Color(indexed=52))
            group_letter.append(dsum.column_letter)

            double = Side(border_style='thin',color="000000")

            for row in dict_sheet[indicator].iter_rows(min_row=1, max_col=stop_i+1, max_row=2):
                for cell in row:
                    cell.border = Border(left=double,right=double,top=double,bottom=double)

            for row in dict_sheet[indicator].iter_rows(min_row=3, max_col=stop_i+1, max_row=len(df)+1):
                for cell in row:
                    cell.border = Border(left=double,right=double)
            
            # Format first 4 id columns (mecontext,cell,siteid,dlchannelbw)
            for i in range(4):
                dict_sheet[indicator].cell(row=2, column=i+1).alignment = Alignment(horizontal='center', vertical='center')

            dict_sheet[indicator].row_dimensions[1].hidden = True
            dict_sheet[indicator].freeze_panes = dict_sheet[indicator]['E3']
            dict_sheet[indicator].sheet_view.zoomScale = 75
        
        ### ------------------ Add CellRelation --------------- ###
        for mo in dict_df:
            if ('CellRelation' in mo) or ('External' in mo):
                ws = wb.create_sheet(title=mo)
                df_concheck = dict_df[mo].configuration
                for r in dataframe_to_rows(df_concheck,index=None,header=True):
                    ws.append(r)

        ######-------------------------------------------------- Adding Summary Sheet ------------------------------------------###############
        print('Calculating Summary')
        ws = wb['Sheet']
        ws.title = 'Summary'

        df_summary = pd.DataFrame(summary_data)

        # ADD FEATURE STATE NAME IN A NEW COlUMN #
        df_summary['Info'] = ''
        df_format = pd.concat([df_format,
                               pd.DataFrame(new_sub_mo_rows)],
                               ignore_index=True)

        if 'FeatureState' in dict_df:
            df_featurestate = dict_df['FeatureState'].configuration.copy()
            df_featurestate['sub_mo'] = 'FeatureState='+df_featurestate['featurestateid']
            dict_featurestate = df_featurestate.set_index('sub_mo').to_dict()['description']

            df_summary.loc[df_summary['Parameter Indicator'].str.contains('FeatureState'),'Info'] = df_summary['Parameter Indicator'].str.split('.').str[0].map(dict_featurestate)

        #Add Action
        df_summary['Action'] = df_summary['Parameter Indicator'].map(df_format.set_index('MO.Parameter').to_dict()['Action'])
        df_summary['Action'] = df_summary['Action'].str.replace('None',"")
        #Add Remark
        df_summary['Remark'] = df_summary['Parameter Indicator'].map(df_format.set_index('MO.Parameter').to_dict()['Remark'])
        df_summary['Remark'] = df_summary['Remark'].str.replace('None',"")
        #Add Rules
        df_summary['Rules'] = df_summary['Parameter Indicator'].map(df_format.set_index('MO.Parameter').to_dict()['Rules'])
        df_summary['Rules'] = df_summary['Rules'].astype(str).str.replace('None',"")

        for r in dataframe_to_rows(df_summary,index=None,header=True):
            ws.append(r)

        header_color = PatternFill(patternType='solid', fgColor=Color(indexed=32))
        summary_color = PatternFill(patternType='solid', fgColor=Color(indexed=48))
        mo_color = PatternFill(patternType='solid', fgColor=Color(indexed=44))

        header_col_index = [0,1,2,3]
        for color_index in header_col_index:
            ws[1][color_index].fill = header_color
            ws[1][color_index].font = Font(color=Color(indexed=1))
            for i in summary_index:
                ws[i][color_index].fill = summary_color
            for i in group_summary_index:
                ws[i][color_index].fill = mo_color

        wb.save(file_path)
        print('Done')

    def create_summary_dict(self,series):
        summary_dict = series.value_counts().to_dict()
        if not ('OK' in summary_dict):
            summary_dict['OK'] = 0
        if not ('NOK' in summary_dict):
            summary_dict['NOK'] = 0
        if not ('NA' in summary_dict):
            summary_dict['NA'] = 0
        
        summary_dict['NA'] = len(series)-summary_dict['OK']-summary_dict['NOK']
        
        return summary_dict
    

class Auditor:
    def __init__(self):
        pass

    def audit(
            self, 
            reference, 
            dict_df
    ):
        '''
        reference : config_reference object
        dict_df   : dictionary containing the CM's to be evaluated
        '''
        global_config_path = os.path.join(os.path.dirname(__file__), 'config.json')
        global_config    = GlobalConfig(global_config_path)
        dict_band = global_config.get_parameter('dict_band')

        dict_concheck = {}
        for mo in reference.moList :
            
            mo_id = mo.lower()+'id'
            
            sub_ids = ['']
            if len(reference.moList[mo][mo_id])>0:
                sub_ids = reference.moList[mo][mo_id]

            for sub_id in sub_ids:
                
                if sub_id != '':
                    sub_mo = mo+'='+sub_id
                else:
                    sub_mo = mo

                print(sub_mo)

                df_settings_mo = reference.settings.loc[(reference.settings['MO']==sub_mo)]
                
                dict_concheck[sub_mo] = dict_df[mo].configuration
                
                if '=' in sub_mo:
                    dict_concheck[sub_mo] = dict_concheck[sub_mo].loc[dict_concheck[sub_mo][mo_id]==sub_id]

                if len(dict_concheck[sub_mo])>0:
                    for index, row in df_settings_mo.iterrows():
                    
                        param = row['Parameter']
                        print('   %s'%param)
                        dependency = row['Dependency']

                        if '{' in dependency:
                            dict_dependency = self.__to_dict_str(dependency)
                            dict_dependency = {value.lower(): key for key, values in dict_dependency.items() for value in values}

                        ########################## Map cell band ###########################
                        #### ini harus ada solusinya kalau pindah operator ga bisa handle dengan logic ini

                        # dict_band = {'T' : 'L900',
                        #             'L' : 'L1800',
                        #             'R' : 'L2100',
                        #             'E' : 'L2300_20',
                        #             'F' : 'L2300_20',
                        #             'V' : 'L2300_10' }
                        
                        dict_concheck[sub_mo]['band'] = np.nan
                        band_pattern = r'[A-Za-z]{3}\d{3}[A-Za-z](.)'
                        dict_concheck[sub_mo]['band_label'] = np.nan
                        if 'eutrancellfddid' in dict_concheck[sub_mo].columns:
                            #dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band']), 'band'] = dict_concheck[sub_mo]['eutrancellfddid'].str[7].map(dict_band)
                            empty_band_label =  dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band_label'])]
                            empty_band_label['band_label'] = empty_band_label['eutrancellfddid'].str.extract(r'[A-Za-z]{3}\d{3}[A-Za-z](.)')
                            #empty_band_label.loc[pd.isna(empty_band_label['band']), 'band'] = empty_band_label['band_label'].map(dict_band)
                            dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band_label'])] = empty_band_label
                        if 'eutrancelltddid' in dict_concheck[sub_mo].columns:
                            #dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band']), 'band'] = dict_concheck[sub_mo]['eutrancelltddid'].str[7].map(dict_band)
                            empty_band_label =  dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band_label'])]
                            empty_band_label['band_label'] = empty_band_label['eutrancelltddid'].str.extract(r'[A-Za-z]{3}\d{3}[A-Za-z](.)')
                            #empty_band_label.loc[pd.isna(empty_band_label['band']), 'band'] = empty_band_label['band_label'].map(dict_band)
                            dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band_label'])] = empty_band_label
                        if 'nrcellduid' in dict_concheck[sub_mo].columns:
                            empty_band_label =  dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band_label'])]
                            empty_band_label['band_label'] = empty_band_label['nrcellduid'].str.extract(r'[A-Za-z]{3}\d{3}[A-Za-z](.)')
                            #empty_band_label.loc[pd.isna(empty_band_label['band']), 'band'] = empty_band_label['band_label'].map(dict_band)
                            dict_concheck[sub_mo].loc[pd.isna(dict_concheck[sub_mo]['band_label'])] = empty_band_label
                        
                        dict_concheck[sub_mo]['band'] = dict_concheck[sub_mo]['band_label'].map(dict_band)

                            #dict_concheck[sub_mo] = dict_concheck[sub_mo].drop()
                        
                        if ('eutrancellfddid' not in dict_concheck[sub_mo].columns) and ('eutrancelltddid' not in dict_concheck[sub_mo].columns) and ('nrcellduid' not in dict_concheck[sub_mo].columns) :
                            # for site level mo, which band related to a site is irrelevant
                            dict_concheck[sub_mo]['band'] = 'L900'
                        #####################################################################
                        
                        result = pd.DataFrame()
                        for band, group in dict_concheck[sub_mo].groupby('band'):

                            if '{' in dependency:
                                band_dependency = dict_dependency[band.lower()]
                            else:
                                band_dependency = dependency

                            #print('%s, %s, %s'%(band,len(group),len(result)))
                            target = row[band]
                            group = self.__audit_param(group,dict_df,param,band_dependency,target)
                            #print(f"Group : \n {group}")
                            #print('    %s, %s, %s'%(band,len(group),len(result)))
                            #print(f"Result : \n {result}")
                            if len(result)>0:
                                result = pd.concat([result,group],ignore_index=True)
                            else:
                                result = group
                        dict_concheck[sub_mo] = result
                        #print(len(dict_concheck[sub_mo]))
        
        return auditResult(dict_concheck,reference)

    
    def __audit_param(self,df_config,dict_df,param,dependency,target):
        ''' Comparing a parameter value in df_config against the corresponding reference value in config_ref '''
        
        dict_rules = {}
        #print(df_config.columns)
        #print("   target : %s"%target)
        #remove all whitespace from target
        if type(target) == str :
           target = re.sub(r'\s+','',target)
           #print("   string target")
        if (str(target)=='*'):
            #print("   wildcard target")
            df_config[param+'_ref'] = df_config[param]
            df_config.loc[pd.notnull(df_config[param]), param+'_check'] = 'OK'
        else:
            ######### check if param dependent ##########
            if dependency != 'None':
                #print("   dependent target")
                dependency = dependency.split(';')
                for dep_param in dependency:
                    ################ Disect Dependency Column ###################
                    
                    #print("   dependency : %s"%dep_param)
                    dep_opr = ""
                    if len(dep_param.split("."))>2:
                        dep_opr = dep_param.split(".")[2]
                    
                    dep_mo = dep_param.split('.')[0] 
                    dep_param = dep_param.split('.')[1]                    
                    dep_mo_id = ''
                    #print('Dep param: %s'%dep_param)

                    if '=' in dep_mo:
                        dep_mo_id = dep_mo.split('=')[1]
                        dep_mo = dep_mo.split('=')[0]
                        
                    df_dependent = dict_df[dep_mo].configuration
                    if dep_mo_id != '':
                        df_dependent = df_dependent.loc[df_dependent[dep_mo.lower()+'id'] == dep_mo_id]

                    mapping_columns = []
                    valid_map_columns = ['mecontext','eutrancellfddid','eutrancelltddid','nrcellduid']
                    for col in df_config.columns:
                        if (col in valid_map_columns) and (col in df_dependent):
                            mapping_columns.append(col)
                    
                    #print(df_config.columns)
                    #print(df_dependent.set_index(mapping_columns))
                    dict_dependency_param = df_dependent.set_index(mapping_columns).to_dict()[dep_param]
                    
                    dep_param = dep_mo + '.' + dep_param
                    if dep_param.split('.')[1] not in df_config:
                        df_config[dep_param] = df_config.set_index(mapping_columns).index.map(dict_dependency_param)
                    else:
                        df_config[dep_param] = df_config[dep_param.split('.')[1]]

                    #print(dict_dependency_param)
                    #print(df_config)
                    
                    ####################################################################################################
                    
                    ### Check if dependency affect target or source value ###
                    if dep_opr != "":
                        ## Dependency Affect Source ##
                        ##### Change source_eq = source operator dep_param #####
                        if dep_opr == "+":
                           df_config[param+'_eq'] = df_config[param] + df_config[dep_param]
                        if dep_opr == "-":
                           df_config[param+'_eq'] = df_config[param] - df_config[dep_param]
                        if dep_opr == "/":
                           df_config[param+'_eq'] = df_config[param] / df_config[dep_param]
                        if dep_opr == "*":
                           df_config[param+'_eq'] = df_config[param] * df_config[dep_param]
                         
                    else:
                        ## Dependency affect target ##

                        if target == '=':
                            df_config[param+'_ref'] = df_config[dep_param]
                        ## add other operator clause here
                        else:
                            ##### Map target reference from dependency #####
                            
                            dict_rules = self.create_rules(target)
                            #print("   Dict rules: %s"%dict_rules)
                            df_config.loc[pd.isnull(df_config[dep_param]), dep_param] = np.nan
                            df_config[dep_param+'_eq'] = df_config[dep_param]
                            try:
                                df_config[dep_param+'_eq'] = df_config[dep_param+'_eq'].astype(float)
                            except:
                                pass
                            for key in dict_rules.keys():
                                if (type(key)==str):
                                    if (('>' in key) or ('<' in key) or ('!=' in key)):
                                        df_config[dep_param+'_eq'] = self.convert_cond_value(df_config[[dep_param+'_eq']],key)
                                        
                            #print(df_config)
        
                            df_config[param+'_ref'] = df_config[dep_param+'_eq'].map(dict_rules)
                            df_config.loc[df_config[param+'_ref'] == '*', param+'_ref'] = df_config[param]
                
                #exact_target = True
                # if (not(is_numeric_dtype(df_config[param+'_ref']))) and (param != 'advcellsupaction'):
                #     df_config[param] = df_config[param].astype(str).str.replace(' ','').str.replace('.0','')
                if param+'_ref' in df_config.columns:
                    df_config.reset_index().loc[df_config.reset_index()[param+'_ref'].apply(lambda x: isinstance(x, str)), param] = df_config.reset_index()[param].astype(str).str.replace(' ','').str.replace('.0','')
            
            if not (param+'_ref' in df_config.columns):
                df_config[param+'_ref'] = target
            
            ######### Split Ref and Basic Types Checking Target ##########
            df_config_check = pd.DataFrame()
            for target, df_config in df_config.groupby(param+'_ref'):
                exact_target = True
                
                if param+'_eq' not in df_config.columns:
                    df_config[param+'_eq'] = df_config[param]
                
                if (type(target)==str):
                    if (('>' in target) or ('<' in target) or ('!' in target) or ('=' in target)):
                        exact_target = False
                    if ('[' in target):
                        list_target = [x for x in target.strip('[]').split(',')]
                        df_config[param+'_eq'] = df_config[param+'_eq'].astype(str).str.replace(' ','').str.replace('\.0','')
                        df_config[param+'_ref'] = target
                        df_config.loc[df_config[param+'_eq'].isin(list_target), param+'_check'] = 'OK'
                        df_config.loc[(pd.notnull(df_config[param+'_eq'])) & (pd.isnull(df_config[param+'_check'])), param+'_check'] = 'NOK'
                        
                    if ('(' in target):
                        target = target.strip("()")
                        lower_bound = target.split(',')[0]
                        upper_bound = target.split(',')[1]
                        
                        if lower_bound.replace("-","").isnumeric() and upper_bound.replace("-","").isnumeric():
                           df_config[param+'_ref'] = f"{lower_bound} to {upper_bound}"
                           lower_bound = int(lower_bound)
                           upper_bound = int(upper_bound)+1
                           
                           if lower_bound>upper_bound:
                              raise(ValueError(f"Error checking parameter {param} with target range {target}. Lower bound larger than upper bound"))
                           
                           list_target = list(range(lower_bound,upper_bound))
                           #print(list_target)
                           df_config.loc[df_config[param+'_eq'].isin(list_target), param+'_check'] = 'OK'
                           #print(df_config)
                           df_config.loc[(pd.notnull(df_config[param+'_eq'])) & (pd.isnull(df_config[param+'_check'])), param+'_check'] = 'NOK'
                        else:
                           raise(ValueError(f"Error checking parameter {param} with target range {target}. Invalid reference target. Range checking only take integer values."))
        
                if (exact_target):
                    try:
                       df_config.loc[pd.isnull(df_config[param+'_eq']), param+'_check'] = 'NA'
                       df_config.loc[(pd.isnull(df_config[param+'_check'])) & (df_config[param+'_eq'] != df_config[param+'_ref']), param+'_check'] = 'NOK' 
                       df_config.loc[(pd.isnull(df_config[param+'_check'])) & (df_config[param+'_eq'] == df_config[param+'_ref']), param+'_check'] = 'OK'
                    except:
                       print('df_config length : %s'%len(df_config))
                else:
                    #print("not exact target")
                    #print(param)
                    df_config[param+'_eq'] = self.convert_cond_value(df_config[[param+'_eq']],target)
                    df_config.loc[df_config[param+'_eq'] == df_config[param+'_ref'], param+'_check'] = 'OK'
                    df_config.loc[df_config[param+'_eq'] != df_config[param+'_ref'], param+'_check'] = 'NOK'
                    
                if len(df_config_check)>0:
                    df_config_check = pd.concat([df_config_check,df_config])
                else:
                    df_config_check = df_config
            df_config = df_config_check
        #print(df_config)
        return df_config
    
    def create_rules(self,target):
        dict_rules = {}
        for rule in target.strip('{}').replace(' ','').split(';'):
            rule = rule.replace(' ','')
            if (rule != '='):
                rule_key = rule.split(':')[0]
                rule_value = rule.split(':')[1]
                if rule_key.replace('-','').isnumeric():
                    rule_key = float(rule_key)
                if rule_value.replace('-','').isnumeric():
                    rule_value = float(rule_value)
                dict_rules[rule_key] = rule_value
        return dict_rules
    
    def convert_cond_value(self,df, target):
        #print(target)
        #print(df)
        operator = re.search("([<>=!:\*]+)", target).group(1)
        operand = re.search("([A-Za-z]+|-?\d+)", target).group()
    
        param = df.columns[0]
    
        if is_numeric_dtype(df[param]):
            df[param] = df[param].astype(float)
        if operand.replace('-','').isnumeric():
            operand = float(operand)
        
        #print(f"{operand} and {operator}")
    
        if operator == '>':
            df.loc[df[param].apply(lambda x: not isinstance(x, str) and x > operand), param] = operator+str(operand).split('.')[0]
        elif operator == '<':
            #print(f"df config before : {df}")
            df.loc[df[param].apply(lambda x: not isinstance(x, str) and x < operand), param] = operator+str(operand).split('.')[0]
            #print(f" operator : {operator}")
            #print(f"df config : {df}")
        elif operator == '!=':
            df.loc[df[param]!=operand, param] = operator+str(operand).split('.')[0]
        elif operator == '>=':
            df.loc[df[param].apply(lambda x: not isinstance(x, str) and x >= operand), param] = operator+str(operand).split('.')[0]
        elif operator == '<=':
            df.loc[df[param].apply(lambda x: not isinstance(x, str) and x <= operand), param] = operator+str(operand).split('.')[0]
    
        return df[param]
    
    def map_param(self,dep_param,df_config,dict_df):
        dep_mo = dep_param.split('.')[0] 
        dep_param = dep_param.split('.')[1]
        
        if not(dep_param in df_config.columns):
            if (dict_df[dep_mo].index.nlevels == 2) and (df_config.index.nlevels==1):
                df_dependent = dict_df[dep_mo].reset_index(1)
            else:
                df_dependent = dict_df[dep_mo]    
            dict_dependency_param = df_dependent.to_dict()[dep_param]
            if (dict_df[dep_mo].index.nlevels == 1) and (df_config.index.nlevels==2):
                df_config[dep_param] = df_config.reset_index(1).index.map(dict_dependency_param)
            else:
                df_config[dep_param] = df_config.index.map(dict_dependency_param)
        
        return df_config
    
    # Function to add double quotes around text before colon and items in lists
    def __to_dict_str(self, text):
        def add_quotes(x):
            key = x.group(1)
            items = x.group(2).split(',')
            quoted_items = ['"' + item.strip() + '"' for item in items]
            return f'"{key.strip()}":[' + ','.join(quoted_items) + ']'
    
        # Use regex to add double quotes around text before colon and items in lists
        valid_dict_string = re.sub(r'([^:,]+):\[([^\[\]]*)\]', lambda x: add_quotes(x), text)

        valid_dict_string = valid_dict_string.replace("{","").replace("}","")
        valid_dict_string = "{" + valid_dict_string + "}"

        # Now, using ast.literal_eval to convert the string to a dictionary
        try:
            dict_dependency = ast.literal_eval(valid_dict_string)
            return dict_dependency
        except (SyntaxError, ValueError):
            print("Error: The input string is not a valid dictionary.")
            return None

    